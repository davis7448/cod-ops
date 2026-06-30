#!/usr/bin/env python3
"""Genera el reporte de reclamo de RECHAZADOS listo para enviar al proveedor.
RECHAZADO lo pone el proveedor (no el cliente) → reclamable. Lee tienda.config.json
y produce un Excel formateado con encabezado, resumen y detalle por pedido.

Uso: python3 gen_reclamo_rechazados.py [ruta/al/tienda.config.json] [YYYY-MM-DD]
"""
import openpyxl, os, sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
CFG_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE, "tienda.config.json")
HOY = sys.argv[2] if len(sys.argv) > 2 else datetime.now().strftime("%Y-%m-%d")
CFG = json.load(open(CFG_PATH, encoding="utf-8"))
COL = CFG["dropi_columnas"]; TIENDA = CFG["tienda"]["nombre"]; MON = CFG["tienda"].get("moneda", "COP")
REPS = CFG["reportes"]["dropi"]

# (clave lógica, encabezado visible) — usa el mapeo del config donde aplica
FIELDS = [("fecha", "Fecha"), ("guia", "N° Guía"), ("transportadora", "Transportadora"),
          ("_cliente", "Cliente"), ("telefono", "Teléfono"), ("departamento", "Departamento"),
          ("ciudad", "Ciudad"), ("_dir", "Dirección"), ("producto", "Producto"),
          ("_cant", "Cant."), ("total", "Valor"), ("_pedido", "N° pedido tienda")]
EXTRA = {"_cliente": "NOMBRE CLIENTE", "_dir": "DIRECCION", "_cant": "CANTIDAD", "_pedido": "NUMERO DE PEDIDO DE TIENDA"}
def colname(key): return EXTRA.get(key) or COL.get(key)

seen = set(); rech = []
for f in REPS:
    if not os.path.exists(f): continue
    ws = openpyxl.load_workbook(f, data_only=True).active
    rows = list(ws.iter_rows(values_only=True)); idx = {h: i for i, h in enumerate(rows[0])}
    def g(r, cn): return r[idx[cn]] if cn in idx else ""
    for r in rows[1:]:
        oid = g(r, COL["id"])
        if oid in seen: continue
        seen.add(oid)
        if str(g(r, COL["estatus"])) != "RECHAZADO": continue
        rech.append([g(r, colname(k)) for k, _ in FIELDS])
ti = [k for k, _ in FIELDS].index("total"); gi = [k for k, _ in FIELDS].index("guia")
tri = [k for k, _ in FIELDS].index("transportadora")
rech.sort(key=lambda x: (str(x[tri] or ""), str(x[0] or "")))
total = sum(float(x[ti] or 0) for x in rech)

wb = Workbook(); ws = wb.active; ws.title = "Reclamo Rechazados"
azul = PatternFill("solid", fgColor="1F4E79"); gris = PatternFill("solid", fgColor="D9E1F2")
bold = Font(bold=True, color="FFFFFF"); boldk = Font(bold=True)
thin = Side(style="thin", color="BBBBBB"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.merge_cells("A1:E1"); ws["A1"] = "RECLAMO — PEDIDOS RECHAZADOS POR EL PROVEEDOR"; ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Tienda: " + TIENDA; ws["A3"] = "Fecha del reclamo: " + HOY
ws["A4"] = "Total pedidos reclamados: %d" % len(rech); ws["A4"].font = boldk
ws["A5"] = "Valor total a reclamar: $%s %s" % (format(round(total), ","), MON); ws["A5"].font = boldk
ws.merge_cells("A7:M7")
ws["A7"] = ("Motivo: estos pedidos fueron marcados RECHAZADO por el proveedor (no por el cliente ni anulados "
            "por nosotros). Causa: sin stock o guía con error. Los que no tienen número de guía nunca se "
            "despacharon. Se solicita revisión y reconocimiento.")
ws["A7"].alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[7].height = 45

hdr = ["#"] + [h for _, h in FIELDS] + ["Motivo probable"]
hr = 9
for j, h in enumerate(hdr, 1):
    c = ws.cell(hr, j, h); c.fill = azul; c.font = bold; c.border = bd
    c.alignment = Alignment(horizontal="center", wrap_text=True)
for i, row in enumerate(rech, 1):
    guia = row[gi]
    motivo = "Sin stock / guía no generada" if guia in (None, "") else "Guía con error (no despachada)"
    disp = [guia or "—" if k == "guia" else row[n] for n, (k, _) in enumerate(FIELDS)]
    vals = [i] + disp + [motivo]
    for j, v in enumerate(vals, 1):
        c = ws.cell(hr + i, j, v); c.border = bd
        if j == ti + 2: c.number_format = "#,##0"; c.font = boldk
        if i % 2 == 0: c.fill = gris
ws.freeze_panes = "A10"
for j, w in enumerate([4, 11, 14, 16, 20, 13, 15, 13, 28, 16, 6, 13, 18, 26], 1):
    ws.column_dimensions[get_column_letter(j)].width = w
out = os.path.join(os.path.dirname(CFG_PATH), "RECLAMO_rechazados_%s.xlsx" % HOY[:7])
wb.save(out)
print("OK ->", out)
print("%d pedidos | $%s a reclamar" % (len(rech), format(round(total), ",")))
